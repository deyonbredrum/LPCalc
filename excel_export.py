# -*- coding: utf-8 -*-
"""
Excel 计算书导出模块 - 完整计算过程版
"""

import io
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def get_lp_conclusion_text(level_text, N, building_attr="", attr_type=""):
    """根据防雷等级和建筑属性生成规范依据文本"""
    if level_text == "一类":
        clause = "第3.0.2条"
        detail = f"制造、使用或贮存爆炸危险物质，且年平均雷击次数N={N:.6f}次/a大于0.05次/a"
    elif level_text == "二类":
        if attr_type in ["crowded", "important", "heritage", "office", "comm", "station", "airport"]:
            clause = "第3.0.3条"
            detail = f"人员密集的公共建筑物或重要场所，且年平均雷击次数N={N:.6f}次/a大于0.05次/a"
        else:
            clause = "第3.0.3条"
            detail = f"一般民用建筑，年平均雷击次数N={N:.6f}次/a大于0.25次/a"
    elif level_text == "三类":
        if attr_type in ["crowded", "important", "heritage", "office", "comm", "station", "airport"]:
            clause = "第3.0.4条"
            detail = f"人员密集的公共建筑物或重要场所，年平均雷击次数N={N:.6f}次/a在0.01~0.05次/a之间"
        else:
            clause = "第3.0.4条"
            detail = f"一般民用建筑，年平均雷击次数N={N:.6f}次/a在0.05~0.25次/a之间"
    else:
        clause = "第3.0.4条"
        detail = f"年平均雷击次数N={N:.6f}次/a小于0.01次/a，可不设防雷"

    return f"根据GB 50057-2010《建筑物防雷设计规范》{clause}规定，{detail}，应判定为{level_text}防雷建筑物。"


def get_ep_conclusion_text(ep_level_text, E):
    """根据电子防护等级生成规范依据文本"""
    if ep_level_text == "A级":
        detail = f"雷电防护等级为A级，拦截效率E={E:.4f} > 0.98"
    elif ep_level_text == "B级":
        detail = f"雷电防护等级为B级，拦截效率0.95 < E={E:.4f} ≤ 0.98"
    elif ep_level_text == "C级":
        detail = f"雷电防护等级为C级，拦截效率0.90 < E={E:.4f} ≤ 0.95"
    elif ep_level_text == "D级":
        detail = f"雷电防护等级为D级，拦截效率0.80 < E={E:.4f} ≤ 0.90"
    else:
        detail = f"雷电防护等级低于D级，拦截效率E={E:.4f} ≤ 0.80，可不设防护"

    return f"根据GB 50343-2012《建筑物电子信息系统防雷技术规范》第5.4.1条规定，{detail}，应判定为{ep_level_text}。"


def format_formula_with_values(formula, values):
    """将公式中的变量替换为实际数值，显示代入过程"""
    result = formula
    for key, val in values.items():
        result = result.replace(key, str(val))
    return result


def export_excel_report(lp_result, ep_result, building_params,
                        cable_params, c_factors, level_text, ep_level_text,
                        export_mode='combined', has_ep=True,
                        building_attr="", attr_type=""):
    """
    导出 Excel 计算书 - 完整计算过程版
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 薄边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== Sheet 1: 摘要（简化版） =====
    ws1 = wb.create_sheet("摘要", 0)

    if has_ep and ep_result:
        title_text = '建筑物防雷及电子信息防护等级计算书'
    else:
        title_text = '建筑物防雷等级计算书'

    ws1.merge_cells('A1:F1')
    cell = ws1['A1']
    cell.value = title_text
    cell.font = Font(name='黑体', size=18, bold=True, color='003366')
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('A2:F2')
    cell = ws1['A2']
    cell.value = f'计算日期：{datetime.now().strftime("%Y年%m月%d日")}'
    cell.font = Font(name='宋体', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws1.row_dimensions[3].height = 10

    row = 4
    ws1.merge_cells(f'A{row}:F{row}')
    cell = ws1[f'A{row}']
    cell.value = '一、建筑物参数'
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    for key, value in building_params.items():
        ws1[f'A{row}'] = key
        ws1[f'B{row}'] = str(value)
        row += 1

    row += 1
    ws1.merge_cells(f'A{row}:F{row}')
    cell = ws1[f'A{row}']
    cell.value = '二、防雷等级计算结果'
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    N = lp_result.get('N', 0)
    ws1[f'A{row}'] = '年预计雷击次数 N'
    ws1[f'B{row}'] = f"{N:.6f} 次/a"
    row += 1

    ws1[f'A{row}'] = '防雷等级'
    ws1[f'B{row}'] = level_text
    ws1[f'B{row}'].font = Font(name='宋体', size=12, bold=True, color='003366')
    row += 1

    ws1[f'A{row}'] = '判断依据'
    conclusion = get_lp_conclusion_text(level_text, N, building_attr, attr_type)
    ws1.merge_cells(f'B{row}:F{row}')
    ws1[f'B{row}'] = conclusion
    ws1[f'B{row}'].font = Font(name='宋体', size=10)
    ws1[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws1.row_dimensions[row].height = 30
    row += 1

    if has_ep and ep_result:
        row += 1
        ws1.merge_cells(f'A{row}:F{row}')
        cell = ws1[f'A{row}']
        cell.value = '三、电子信息防护等级计算结果'
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

        ws1[f'A{row}'] = '总年预计雷击次数 N'
        ws1[f'B{row}'] = f"{ep_result.get('N', 0):.6f} 次/a"
        row += 1

        E = ep_result.get('E', 0)
        ws1[f'A{row}'] = '拦截效率 E'
        ws1[f'B{row}'] = f"{E:.4f}"
        row += 1

        ws1[f'A{row}'] = '防护等级'
        ws1[f'B{row}'] = ep_level_text
        color_map = {'A级': 'CC0000', 'B级': 'FF8800', 'C级': '0066CC', 'D级': '008800'}
        ws1[f'B{row}'].font = Font(name='宋体', size=12, bold=True, color=color_map.get(ep_level_text, '000000'))
        row += 1

        ws1[f'A{row}'] = '判断依据'
        ep_conclusion = get_ep_conclusion_text(ep_level_text, E)
        ws1.merge_cells(f'B{row}:F{row}')
        ws1[f'B{row}'] = ep_conclusion
        ws1[f'B{row}'].font = Font(name='宋体', size=10)
        ws1[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws1.row_dimensions[row].height = 30

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws1.column_dimensions[col].width = 20

    # ===== Sheet 2: 防雷等级详细计算过程 =====
    ws2 = wb.create_sheet("防雷等级详细计算")

    row = 1
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.value = '防雷等级详细计算过程'
    cell.font = Font(name='黑体', size=16, bold=True, color='003366')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # 获取计算数据
    Td = lp_result.get('Td', 0)
    Ng = lp_result.get('Ng', 0)
    D = lp_result.get('D', 0)
    Ae = lp_result.get('Ae', 0)
    k = lp_result.get('k', 0)
    N = lp_result.get('N', 0)
    city = lp_result.get('city', '')
    L = lp_result.get('L', 0)
    W = lp_result.get('W', 0)
    H = lp_result.get('H', 0)
    surrounding_type = lp_result.get('surrounding_type', '')
    Lz = lp_result.get('Lz', 0)

    # 计算步骤1: 年雷暴日
    ws2[f'A{row}'] = '步骤1：确定年雷暴日'
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    row += 1

    ws2[f'A{row}'] = '依据'
    ws2[f'B{row}'] = f'GB 50057-2010 附录A，查表得 {city} 年雷暴日'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算公式'
    ws2[f'B{row}'] = 'Td = 查表值'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '代入数据'
    ws2[f'B{row}'] = f'Td = {Td:.2f} d/a'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算结果'
    ws2[f'B{row}'] = f'Td = {Td:.2f} d/a'
    ws2.merge_cells(f'B{row}:E{row}')
    ws2[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
    row += 2

    # 计算步骤2: 雷击大地密度
    ws2[f'A{row}'] = '步骤2：计算雷击大地密度'
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    row += 1

    ws2[f'A{row}'] = '依据'
    ws2[f'B{row}'] = 'GB 50057-2010 公式（附A.4）'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算公式'
    ws2[f'B{row}'] = 'Ng = 0.1 × Td'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '代入数据'
    ws2[f'B{row}'] = f'Ng = 0.1 × {Td:.2f}'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算结果'
    ws2[f'B{row}'] = f'Ng = {Ng:.4f} 次/(km²·a)'
    ws2.merge_cells(f'B{row}:E{row}')
    ws2[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
    row += 2

    # 计算步骤3: 扩大宽度
    ws2[f'A{row}'] = '步骤3：计算扩大宽度'
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    row += 1

    ws2[f'A{row}'] = '依据'
    ws2[f'B{row}'] = 'GB 50057-2010 附录A'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算公式'
    ws2[f'B{row}'] = 'D = √(H × (200 - H))'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '代入数据'
    ws2[f'B{row}'] = f'D = √({H:.1f} × (200 - {H:.1f})) = √({H:.1f} × {200 - H:.1f})'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算结果'
    ws2[f'B{row}'] = f'D = {D:.3f} m'
    ws2.merge_cells(f'B{row}:E{row}')
    ws2[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
    row += 2

    # 计算步骤4: 等效面积
    ws2[f'A{row}'] = '步骤4：计算等效面积'
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    row += 1

    ws2[f'A{row}'] = '依据'
    ws2[f'B{row}'] = 'GB 50057-2010 附录A'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '周边情况'
    ws2[f'B{row}'] = surrounding_type
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    # 根据周边情况显示不同的公式
    if surrounding_type == "周边2D范围内无其他建筑":
        formula = "Ae = LW + 2(L+W)D + πD²"
        formula_with_values = f"Ae = {L}×{W} + 2×({L}+{W})×{D:.3f} + π×{D:.3f}²"
    elif surrounding_type == "周边2D范围内有等高或比它低的其他建筑物，但不在所考虑建筑物以外100m的保护范围内":
        formula = "Ae = LW + 2(L+W)D + πD² - Lz×D/2"
        formula_with_values = f"Ae = {L}×{W} + 2×({L}+{W})×{D:.3f} + π×{D:.3f}² - {Lz}×{D:.3f}/2"
    elif surrounding_type == "周边2D范围内有等高或比它低的其他建筑物":
        formula = "Ae = LW + (L+W)D + πD²/4"
        formula_with_values = f"Ae = {L}×{W} + ({L}+{W})×{D:.3f} + π×{D:.3f}²/4"
    elif surrounding_type == "周边2D范围内有比它高的其他建筑物":
        formula = "Ae = LW + 2(L+W)D + πD² - Lz×D"
        formula_with_values = f"Ae = {L}×{W} + 2×({L}+{W})×{D:.3f} + π×{D:.3f}² - {Lz}×{D:.3f}"
    elif surrounding_type == "周边2D范围内都有比它高的其他建筑物":
        formula = "Ae = LW"
        formula_with_values = f"Ae = {L}×{W}"
    else:
        formula = "Ae = LW + 2(L+W)D + πD²"
        formula_with_values = f"Ae = {L}×{W} + 2×({L}+{W})×{D:.3f} + π×{D:.3f}²"

    ws2[f'A{row}'] = '计算公式'
    ws2[f'B{row}'] = formula
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '代入数据'
    ws2[f'B{row}'] = formula_with_values
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算结果'
    ws2[f'B{row}'] = f'Ae = {Ae:.6f} km²'
    ws2.merge_cells(f'B{row}:E{row}')
    ws2[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
    row += 2

    # 计算步骤5: 校正系数
    ws2[f'A{row}'] = '步骤5：确定校正系数'
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    row += 1

    ws2[f'A{row}'] = '依据'
    ws2[f'B{row}'] = f'GB 50057-2010 第4.2.1条，{building_params.get("建筑物状况", "一般情况")}'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算公式'
    ws2[f'B{row}'] = 'k = 查表值'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '代入数据'
    ws2[f'B{row}'] = f'k = {k:.2f}'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算结果'
    ws2[f'B{row}'] = f'k = {k:.2f}'
    ws2.merge_cells(f'B{row}:E{row}')
    ws2[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
    row += 2

    # 计算步骤6: 年预计雷击次数
    ws2[f'A{row}'] = '步骤6：计算年预计雷击次数'
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    row += 1

    ws2[f'A{row}'] = '依据'
    ws2[f'B{row}'] = 'GB 50057-2010 公式（附A.2）'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算公式'
    ws2[f'B{row}'] = 'N = k × Ng × Ae'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '代入数据'
    ws2[f'B{row}'] = f'N = {k:.2f} × {Ng:.4f} × {Ae:.6f}'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '计算结果'
    ws2[f'B{row}'] = f'N = {N:.6f} 次/a'
    ws2.merge_cells(f'B{row}:E{row}')
    ws2[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
    row += 2

    # 计算步骤7: 防雷等级判定
    ws2[f'A{row}'] = '步骤7：防雷等级判定'
    ws2.merge_cells(f'A{row}:E{row}')
    cell = ws2[f'A{row}']
    cell.font = Font(name='黑体', size=12, bold=True, color='003366')
    row += 1

    ws2[f'A{row}'] = '依据'
    ws2[f'B{row}'] = f'GB 50057-2010 第3.0.2~3.0.4条，建筑物属性：{building_params.get("建筑物属性", "一般民用建筑")}'
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '判定条件'
    if level_text == "一类":
        condition = "N > 0.05 次/a 且 爆炸危险场所"
    elif level_text == "二类":
        condition = "N > 0.05 次/a（人员密集/重要场所）或 N > 0.25 次/a（一般民用）"
    elif level_text == "三类":
        condition = "N > 0.01 次/a（人员密集/重要场所）或 N > 0.05 次/a（一般民用）"
    else:
        condition = "N ≤ 0.01 次/a（人员密集/重要场所）或 N ≤ 0.05 次/a（一般民用）"
    ws2[f'B{row}'] = condition
    ws2.merge_cells(f'B{row}:E{row}')
    row += 1

    ws2[f'A{row}'] = '判定结果'
    ws2[f'B{row}'] = f'防雷等级：{level_text}'
    ws2.merge_cells(f'B{row}:E{row}')
    ws2[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
    row += 2

    # 设置列宽
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = 22

    # ===== Sheet 3: 电子信息防护详细计算过程（仅当 has_ep=True） =====
    if has_ep and ep_result:
        ws3 = wb.create_sheet("电子防护详细计算")

        row = 1
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.value = '电子信息系统防护等级详细计算过程'
        cell.font = Font(name='黑体', size=16, bold=True, color='003366')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2

        # 步骤1: 确定线缆截收面积
        ws3[f'A{row}'] = '步骤1：确定入户设施截收面积'
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        row += 1

        ws3[f'A{row}'] = '依据'
        ws3[f'B{row}'] = 'GB 50343-2012 附录A'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '电源线缆入户方式'
        ws3[f'B{row}'] = cable_params.get('电源线缆入户方式', '未选择')
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '信号线缆入户方式'
        ws3[f'B{row}'] = cable_params.get('信号线缆入户方式', '未选择')
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        Ae1 = ep_result.get('Ae1', 0)
        Ae2 = ep_result.get('Ae2', 0)
        ws3[f'A{row}'] = '截收面积结果'
        ws3[f'B{row}'] = f"Ae1' = {Ae1:.6f} km², Ae2' = {Ae2:.6f} km²"
        ws3.merge_cells(f'B{row}:E{row}')
        ws3[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
        row += 2

        # 步骤2: 计算 N2
        ws3[f'A{row}'] = '步骤2：计算入户设施年预计雷击次数 N2'
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        row += 1

        ws3[f'A{row}'] = '依据'
        ws3[f'B{row}'] = 'GB 50343-2012 附录A'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '计算公式'
        ws3[f'B{row}'] = "N2 = Ng × (Ae1' + Ae2')"
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        N1 = ep_result.get('N1', 0)
        N2 = ep_result.get('N2', 0)
        ws3[f'A{row}'] = '代入数据'
        ws3[f'B{row}'] = f"N2 = {Ng:.4f} × ({Ae1:.6f} + {Ae2:.6f})"
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '计算结果'
        ws3[f'B{row}'] = f"N2 = {N2:.6f} 次/a"
        ws3.merge_cells(f'B{row}:E{row}')
        ws3[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
        row += 2

        # 步骤3: 计算总N
        ws3[f'A{row}'] = '步骤3：计算总年预计雷击次数 N'
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        row += 1

        ws3[f'A{row}'] = '计算公式'
        ws3[f'B{row}'] = 'N = N1 + N2'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        N_total = ep_result.get('N', 0)
        ws3[f'A{row}'] = '代入数据'
        ws3[f'B{row}'] = f"N = {N1:.6f} + {N2:.6f}"
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '计算结果'
        ws3[f'B{row}'] = f"N = {N_total:.6f} 次/a"
        ws3.merge_cells(f'B{row}:E{row}')
        ws3[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
        row += 2

        # 步骤4: C因子明细
        ws3[f'A{row}'] = '步骤4：C因子确定'
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        row += 1

        ws3[f'A{row}'] = '依据'
        ws3[f'B{row}'] = 'GB 50343-2012 表5.4.3-2 ~ 表5.4.3-3'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        # C因子表格
        factor_names = {
            'C1': '建筑物材料结构',
            'C2': '信息系统重要程度',
            'C3': '设备耐冲击能力',
            'C4': '雷电防护区',
            'C5': '雷击事故后果',
            'C6': '区域雷暴等级',
        }
        for key, value in c_factors.items():
            ws3[f'A{row}'] = key
            ws3[f'B{row}'] = factor_names.get(key, '')
            ws3[f'C{row}'] = value.get('type', '')
            ws3[f'D{row}'] = str(value.get('val', 0))
            row += 1

        row += 1
        C_total = ep_result.get('C', 0)
        ws3[f'A{row}'] = 'C因子总和'
        ws3[f'B{row}'] = f"C = C1+C2+C3+C4+C5+C6 = {C_total:.2f}"
        ws3.merge_cells(f'B{row}:E{row}')
        ws3[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
        row += 2

        # 步骤5: 计算Nc
        ws3[f'A{row}'] = '步骤5：计算可接受最大年雷击次数 Nc'
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        row += 1

        ws3[f'A{row}'] = '依据'
        ws3[f'B{row}'] = 'GB 50343-2012 附录A'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '计算公式'
        ws3[f'B{row}'] = 'Nc = 0.58 / C'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        Nc = ep_result.get('Nc', 0)
        ws3[f'A{row}'] = '代入数据'
        ws3[f'B{row}'] = f"Nc = 0.58 / {C_total:.2f}"
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '计算结果'
        ws3[f'B{row}'] = f"Nc = {Nc:.6f} 次/a"
        ws3.merge_cells(f'B{row}:E{row}')
        ws3[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
        row += 2

        # 步骤6: 计算拦截效率E
        ws3[f'A{row}'] = '步骤6：计算拦截效率 E'
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        row += 1

        ws3[f'A{row}'] = '依据'
        ws3[f'B{row}'] = 'GB 50343-2012 附录A'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '计算公式'
        ws3[f'B{row}'] = 'E = 1 - Nc / N'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        E = ep_result.get('E', 0)
        ws3[f'A{row}'] = '代入数据'
        ws3[f'B{row}'] = f"E = 1 - {Nc:.6f} / {N_total:.6f}"
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '计算结果'
        ws3[f'B{row}'] = f"E = {E:.4f}"
        ws3.merge_cells(f'B{row}:E{row}')
        ws3[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')
        row += 2

        # 步骤7: 防护等级判定
        ws3[f'A{row}'] = '步骤7：防护等级判定'
        ws3.merge_cells(f'A{row}:E{row}')
        cell = ws3[f'A{row}']
        cell.font = Font(name='黑体', size=12, bold=True, color='003366')
        row += 1

        ws3[f'A{row}'] = '依据'
        ws3[f'B{row}'] = 'GB 50343-2012 第5.4.1条'
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '判定条件'
        if ep_level_text == "A级":
            condition = "E > 0.98"
        elif ep_level_text == "B级":
            condition = "0.95 < E ≤ 0.98"
        elif ep_level_text == "C级":
            condition = "0.90 < E ≤ 0.95"
        elif ep_level_text == "D级":
            condition = "0.80 < E ≤ 0.90"
        else:
            condition = "E ≤ 0.80"
        ws3[f'B{row}'] = condition
        ws3.merge_cells(f'B{row}:E{row}')
        row += 1

        ws3[f'A{row}'] = '判定结果'
        ws3[f'B{row}'] = f'防护等级：{ep_level_text}'
        ws3.merge_cells(f'B{row}:E{row}')
        ws3[f'B{row}'].font = Font(name='宋体', size=11, bold=True, color='003366')

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws3.column_dimensions[col].width = 22

    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    return file_buffer


def export_excel_separate(lp_result, ep_result, building_params,
                          cable_params, c_factors, level_text, ep_level_text,
                          building_attr="", attr_type=""):
    """分开导出：防雷计算书 + 电子防护计算书"""
    lp_buffer = export_excel_report(
        lp_result=lp_result,
        ep_result=None,
        building_params=building_params,
        cable_params=cable_params,
        c_factors={},
        level_text=level_text,
        ep_level_text=None,
        has_ep=False,
        building_attr=building_attr,
        attr_type=attr_type
    )

    ep_buffer = export_excel_report(
        lp_result=lp_result,
        ep_result=ep_result,
        building_params=building_params,
        cable_params=cable_params,
        c_factors=c_factors,
        level_text=level_text,
        ep_level_text=ep_level_text,
        has_ep=True,
        building_attr=building_attr,
        attr_type=attr_type
    )

    return lp_buffer, ep_buffer